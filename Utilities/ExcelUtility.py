import openpyxl
from openpyxl.styles import PatternFill
import re

from selenium.webdriver import Keys, ActionChains


def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rowNum,columnNum):
    workboook = openpyxl.load_workbook(file)
    sheet = workboook[sheetName]
    return sheet.cell(rowNum, columnNum).value

def writeData(file,sheetName,rowNum,columnNum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowNum,columnNum).value = data
    workbook.save(file)

def fillGreenColor(file,sheetName,rowNum,columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212', end_color='60b212',fill_type='solid')
    sheet.cell(rowNum,columnNum).fill = greenFill
    workbook.save(file)

def fillRedColor(file,sheetName,rowNum,columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(rowNum,columnNum).fill = redFill
    workbook.save(file)


#def extract_price(web_text):      #Takes dirty web text (like '₹49,000') and returns a clean float (49000.0).

    # This regex means: "Keep only digits (\d) and decimals (\.), delete everything else"
  #  clean_string = re.sub(r'[^\d.]', '', web_text)

    # Convert it to a float and return it
   # return float(clean_string)


#def force_clear_and_type(driver, element, text_to_type):  #Highlights all existing text, deletes it, and types new text.
 #   """Uses raw Javascript to bypass UI restrictions and force the box empty."""
    # 1. Reach into the HTML and instantly delete the value
  #  driver.execute_script("arguments[0].value = '';", element)

    # 2. Type directly over the highlighted text
   # element.send_keys(text_to_type)



